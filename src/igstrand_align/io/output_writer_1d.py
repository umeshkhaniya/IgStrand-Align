"""1D alignment workbook writing scaffold."""

from pathlib import Path

from ..alignment.color_rules import STRAND_COLOR_MAP_1D
from ..core.exceptions import DependencyError


class OutputWriter1D:
    """Render a 1D alignment result into an output artifact."""

    def __init__(self):
        self._font_cls = None
        self._pattern_fill_cls = None

    @staticmethod
    def _split_ig_label(number_string: str):
        for index, char in enumerate(number_string):
            if char.isdigit():
                return number_string[:index], number_string[index:]
        return "", number_string

    def _fill_residue_cell(self, cell, residue_mapping):
        if residue_mapping is None:
            cell.value = ""
            return

        if self._font_cls is None or self._pattern_fill_cls is None:
            raise RuntimeError("OpenPyXL styles were not initialized before writing cells.")

        cell.value = residue_mapping.residue_code
        if residue_mapping.ig_label.endswith("50"):
            color_code = "FFD700"
        elif residue_mapping.is_loop:
            color_code = STRAND_COLOR_MAP_1D["loop"]
        else:
            strand_label, _ = self._split_ig_label(residue_mapping.ig_label)
            color_code = STRAND_COLOR_MAP_1D.get(strand_label.strip("+-_"), "FFFFFF")

        cell.fill = self._pattern_fill_cls(
            start_color=color_code,
            end_color=color_code,
            fill_type="solid",
        )
        cell.font = self._font_cls(size=12)

    def write(self, result, destination):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ModuleNotFoundError as exc:
            raise DependencyError(
                "Writing 1D Excel output requires the 'openpyxl' package."
            ) from exc

        self._font_cls = Font
        self._pattern_fill_cls = PatternFill

        workbook = Workbook()
        worksheet = workbook.active

        for column, header in enumerate(result.metadata_headers, start=1):
            worksheet.cell(row=1, column=column, value=header).font = Font(size=14)

        offset = len(result.metadata_headers)
        for column, ig_label in enumerate(result.ig_labels, start=offset + 1):
            worksheet.cell(row=1, column=column, value=ig_label).font = Font(size=14)

        for row, domain in enumerate(result.domains, start=2):
            metadata = [
                domain.structure_id,
                domain.ref_pdb_name,
                domain.tm_score,
                domain.ig_type,
                domain.domain_residue_range,
                domain.ig_residue_range,
                domain.seq_id,
                domain.n_res_aligned,
                ",".join(domain.undefined_residues),
            ]
            for column, value in enumerate(metadata, start=1):
                worksheet.cell(row=row, column=column, value=value)

            for column, ig_label in enumerate(result.ig_labels, start=offset + 1):
                residue_mapping = domain.residue_map.get(ig_label)
                self._fill_residue_cell(worksheet.cell(row=row, column=column), residue_mapping)

        output_path = Path(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
