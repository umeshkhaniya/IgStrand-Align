"""2D alignment workbook writing scaffold."""

from copy import copy
from pathlib import Path

from ..alignment.color_rules import STRAND_COLOR_MAP_2D
from ..core.exceptions import DependencyError, MissingTemplateError


class OutputWriter2D:
    """Render a 2D alignment result into an output artifact."""

    def __init__(self, template_repository, numbering_name: str, template_dimensions):
        self.template_repository = template_repository
        self.numbering_name = numbering_name
        self.template_dimensions = template_dimensions

    @staticmethod
    def _split_ig_label(number_string: str):
        for index, char in enumerate(number_string):
            if char.isdigit():
                return number_string[:index], number_string[index:]
        return "", number_string

    def _template_shape(self, ig_type: str):
        return (
            self.template_dimensions.get(f"{ig_type}_row_range", self.template_dimensions["V_row_range"]),
            self.template_dimensions.get(
                f"{ig_type}_column_range", self.template_dimensions["V_column_range"]
            ),
        )

    def write(self, result, destination):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ModuleNotFoundError as exc:
            raise DependencyError(
                "Writing 2D Excel output requires the 'openpyxl' package."
            ) from exc

        workbook = Workbook()
        worksheet_out = workbook.active
        columns_to_shift = 0

        for panel in result.panels:
            try:
                template_path = self.template_repository.resolve_template_path(
                    panel.template_name,
                    self.numbering_name,
                )
            except MissingTemplateError:
                raise

            template_workbook = load_workbook(template_path)
            worksheet = template_workbook.active
            template_rows, template_cols = self._template_shape(panel.domain.ig_type)

            residue_map_no_strand = {}
            for ig_label, residue_mapping in panel.domain.residue_map.items():
                _, numeric = self._split_ig_label(ig_label)
                residue_map_no_strand[numeric] = residue_mapping

            for row in range(1, template_rows + 1):
                for col in range(template_cols, 0, -1):
                    source_cell = worksheet.cell(row=row, column=col)
                    dest_col = col + columns_to_shift
                    dest_cell = worksheet_out.cell(row=row, column=dest_col)
                    cell_value = source_cell.value
                    residue_mapping = residue_map_no_strand.get(str(cell_value))

                    if residue_mapping is not None:
                        if residue_mapping.is_loop:
                            color_code = STRAND_COLOR_MAP_2D["loop"]
                        elif str(cell_value).endswith("50"):
                            color_code = "FFFF00"
                        else:
                            color_code = STRAND_COLOR_MAP_2D.get(str(cell_value)[0], "FFFFFF")

                        dest_cell.value = residue_mapping.residue_code
                        dest_cell.fill = PatternFill(
                            start_color=color_code,
                            end_color=color_code,
                            fill_type="solid",
                        )
                    else:
                        dest_cell.fill = copy(source_cell.fill)
                        if cell_value == panel.domain.ig_type:
                            dest_cell.value = panel.domain.ig_type + "_" + panel.domain.structure_id
                            ref_cell = worksheet_out.cell(row=row, column=dest_col + 3)
                            ref_cell.value = panel.reference_structure
                            ref_cell.font = Font(bold=True, size=16)
                        elif (
                            cell_value
                            and len(str(cell_value)) == 4
                            and str(cell_value)[0].isdigit()
                        ):
                            dest_cell.value = ""
                        else:
                            dest_cell.value = cell_value

                    dest_cell.alignment = Alignment(horizontal="center")
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.border = copy(source_cell.border)

            columns_to_shift += template_cols

        output_path = Path(destination)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
